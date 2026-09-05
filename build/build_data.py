# -*- coding: utf-8 -*-
"""Build merged COURSES data: autumn from new timetable, spring from existing page."""
import openpyxl, json, re
from collections import Counter, defaultdict

NEW_FILE = '../2026年秋季学期课表.xlsx'
LEGACY_FILE = 'legacy_plan_0828.json'   # 0828 开课计划派生基准（学科归类/校区核对）
SPRING_FILE = '../2026-2027学年秋季和春季开课计划表0903.xlsx'  # 官网 0903 更新版
OUT = 'courses_merged.json'

# ---------- 1. load legacy plan (用于学科归类 map 与校区交叉核对) ----------
existing = json.load(open(LEGACY_FILE, encoding='utf-8'))

# ---------- 2. load new timetable ----------
wb = openpyxl.load_workbook(NEW_FILE, data_only=True)
ws = wb['sheet0']
all_rows = []
for r in range(2, ws.max_row + 1):
    vals = [ws.cell(row=r, column=col).value for col in range(1, 25)]
    if all(v is None for v in vals):
        continue
    all_rows.append(vals)

courses = []   # anchor + slots
cur = None
for vals in all_rows:
    if vals[0] is not None:
        if cur: courses.append(cur)
        cur = {'anchor': vals, 'slots': []}
    else:
        cur['slots'].append(vals)
if cur: courses.append(cur)

# ---------- 3. subjectCode map from existing ----------
sc_map = defaultdict(lambda: defaultdict(Counter))   # sc -> first -> Counter(second)
for c in existing:
    sc_map[c['subjectCode']][c['first']][c['second']] += 1

SPECIAL_SC = {
    '025100': ('应用经济学', '金融学'),
    '035400': ('其他 / 自设学科', '知识产权'),
    '050101': ('中国语言文学', '文艺学'),
    '050102': ('中国语言文学', '语言学及应用语言学'),
    '075100': ('大气科学', '气象学'),
    '125100': ('工商管理', '工商管理'),
    '125200': ('公共管理', '公共管理'),
    '99J100': ('其他 / 自设学科', '人居科学'),
    '99J1X1': ('其他 / 自设学科', '人居前沿及交叉科学'),
    '99J1X2': ('其他 / 自设学科', '广义建筑学'),
}

def derive_first_second(code, sc, disc):
    """Return (first, second) using subjectCode map, falling back to H discipline."""
    if sc in SPECIAL_SC:
        return SPECIAL_SC[sc]
    fs = sc_map.get(sc)
    if fs:
        firsts = sorted(fs.keys(), key=lambda k: -sum(fs[k].values()))
        first = firsts[0]
        seconds = fs[first]
        if disc and disc in seconds:
            return (first, disc)
        if disc and disc == first:
            return (first, '一级学科课程')
        second = max(seconds.items(), key=lambda kv: kv[1])[0]
        return (first, second)
    # fallback: try disc as first
    if disc:
        return (disc, disc if disc in {c['second'] for c in existing} else '一级学科课程')
    return ('其他 / 自设学科', '自设课程')

def campus_from_code(code):
    mm = re.search(r'(?:P|M|D)[A-Z]?\d{3,4}([HYZ])', code)
    if mm: return {'H':'雁栖湖','Y':'玉泉路','Z':'中关村'}[mm.group(1)]
    return None

def parse_slot_m(m, weeks):
    """Parse '周二(5-6)' or '周六(1-3,5-7)' -> list of (day, start, end, label)."""
    mm = re.match(r'周([一二三四五六日天])\((.+)\)', m.strip())
    if not mm: return []
    day = {'一':1,'二':2,'三':3,'四':4,'五':5,'六':6,'日':7,'天':7}[mm.group(1)]
    out = []
    for part in mm.group(2).split(','):
        part = part.strip()
        if '-' in part:
            a, b = part.split('-'); out.append((day, int(a), int(b)))
        else:
            out.append((day, int(part), int(part)))
    return out

DAY_CN = {1:'周一',2:'周二',3:'周三',4:'周四',5:'周五',6:'周六',7:'周日'}

# ---------- 4. build autumn records ----------
existing_by_code = defaultdict(list)
for c in existing: existing_by_code[c['code']].append(c)

autumn = []
warn = []
for i, c in enumerate(courses, 1):
    a = c['anchor']
    code = str(a[2]).strip()
    name = str(a[3]).strip()
    college = str(a[1]).strip()
    disc = str(a[7]).strip() if a[7] else ''
    hours_credits = str(a[8]).strip().split('/')
    hours = float(hours_credits[0])
    credits = float(hours_credits[1])
    category = str(a[5]).strip()
    level = str(a[6]).strip()
    cap = a[9]
    enrolled = a[10] if a[10] is not None else 0
    exam = str(a[15]).strip() if a[15] else ''
    teach = str(a[14]).strip() if a[14] else ''
    room = str(a[13]).strip() if a[13] else ''
    english = str(a[4]).strip() if a[4] else ''

    sc = code[6:12] if len(code) >= 12 else ''
    first, second = derive_first_second(code, sc, disc)

    # campus: from code, cross-check existing
    campus = campus_from_code(code)
    ex_match = None
    for e in existing_by_code.get(code, []):
        if e['college'] == college or not ex_match:
            ex_match = e
    if campus is None:
        campus = ex_match['campus'] if ex_match else '雁栖湖'
    elif ex_match and ex_match['campus'] != campus:
        warn.append(f'campus mismatch {code}: code={campus} existing={ex_match["campus"]}')

    # slots
    slots = []
    slot_rows = [a] + c['slots']
    for sr in slot_rows:
        mval = sr[12]
        wval = sr[11]
        if not mval: continue
        wtext = str(wval).strip() if wval else ''
        for (day, s, e) in parse_slot_m(str(mval), wtext):
            slots.append({'day': day, 'p': f'{s}-{e}', 'start': s, 'end': e, 'w': wtext})

    rec = {
        'id': i,
        'code': code, 'name': name, 'en': english,
        'college': college, 'campus': campus, 'semester': '秋季',
        'category': category, 'discipline': disc, 'subjectCode': sc,
        'first': first, 'second': second,
        'level': level, 'hours': hours, 'credits': credits,
        'capacity': cap, 'enrolled': int(enrolled) if enrolled else 0,
        'exam': exam, 'teach': teach,
        'chief': str(a[16]).strip() if a[16] else '', 'chiefUnit': str(a[17]).strip() if a[17] else '',
        'main': str(a[18]).strip() if a[18] else '', 'mainUnit': str(a[19]).strip() if a[19] else '',
        'ta': str(a[20]).strip() if a[20] else '', 'taUnit': str(a[21]).strip() if a[21] else '',
        'convener': str(a[22]).strip() if a[22] else '',
        'room': room, 'slots': slots,
    }
    autumn.append(rec)

# ---------- 5. load official Hangzhou courses ----------
hangzhou = json.load(open('hangzhou_courses.json', encoding='utf-8'))
bad = [c.get('code', '') for c in hangzhou
       if not c.get('code', '').startswith('280216') or c.get('campus') != '杭州'
       or c.get('semester') not in {'秋季', '春季'}]
if bad:
    raise ValueError(f'invalid Hangzhou course records: {bad[:5]}')
print(f'loaded Hangzhou courses: {len(hangzhou)}')

# ---------- 6. spring records (from 官网 0903 更新版开课计划) ----------
spring = []
legacy_spring = {c['code']: c for c in existing if c['semester'] == '春季'}
wb2 = openpyxl.load_workbook(SPRING_FILE, data_only=True)
ws2 = wb2['2026-2027学年春季学期课程计划情况']
for r in ws2.iter_rows(min_row=3, values_only=True):
    if r[0] is None: continue
    seq, college, code, name, campus, sem, category, discipline, hours, credits = r
    code = str(code).strip(); name = str(name).strip()
    college = str(college).strip(); campus = str(campus).strip()
    discipline = str(discipline).strip() if discipline else ''
    category = str(category).strip() if category else ''
    sc = code[6:12] if len(code) >= 12 else ''
    legacy_rec = legacy_spring.get(code)
    if legacy_rec and legacy_rec.get('first'):
        first, second = legacy_rec['first'], legacy_rec['second']
    else:
        first, second = derive_first_second(code, sc, discipline)
    spring.append({
        'id': 0,  # filled later
        'code': code, 'name': name, 'en': '',
        'college': college, 'campus': campus, 'semester': '春季',
        'category': category, 'discipline': discipline, 'subjectCode': sc,
        'first': first, 'second': second,
        'level': '', 'hours': float(hours), 'credits': float(credits),
        'capacity': None, 'enrolled': 0,
        'exam': '', 'teach': '',
        'chief': '', 'chiefUnit': '', 'main': '', 'mainUnit': '', 'ta': '', 'taUnit': '', 'convener': '',
        'room': '', 'slots': [],
    })
wb2.close()

# 杭州文件可在正式春季课表发布后独立更新；避免与全校计划中的同编码课程重复。
hangzhou_keys = {(c['code'], c['semester']) for c in hangzhou}
spring = [c for c in spring if (c['code'], c['semester']) not in hangzhou_keys]

# ---------- 7. assign ids, validate ----------
all_courses = autumn + spring + hangzhou
for idx, c in enumerate(all_courses, 1):
    c['id'] = idx

# validation
issues = []
autumn_all = autumn + [c for c in hangzhou if c['semester'] == '秋季']
spring_all = spring + [c for c in hangzhou if c['semester'] == '春季']
codes_a = [c['code'] for c in autumn_all]
dup_a = [k for k, v in Counter(codes_a).items() if v > 1]
if dup_a: issues.append(f'duplicate autumn codes: {dup_a[:10]}')
codes_s = [c['code'] for c in spring_all]
dup_s = [k for k, v in Counter(codes_s).items() if v > 1]
if dup_s: issues.append(f'duplicate spring codes: {dup_s[:10]}')
no_slots = [c['code'] for c in autumn_all if not c['slots']]
if no_slots: issues.append(f'autumn courses with no slots: {len(no_slots)} {no_slots[:5]}')
no_room = [c['code'] for c in autumn_all if not c['room']]
if no_room: issues.append(f'autumn courses with no room: {len(no_room)} {no_room[:5]}')
over = [c['code'] for c in autumn_all if c['capacity'] and c['enrolled'] > c['capacity']]
if over: issues.append(f'enrolled>capacity: {len(over)} {over[:5]}')
print('warnings:', len(warn), warn[:10])
print('issues:', issues if issues else 'NONE')
print(f'total: autumn={len(autumn_all)} spring={len(spring_all)} all={len(all_courses)}')

json.dump(all_courses, open(OUT, 'w', encoding='utf-8'), ensure_ascii=False, separators=(',', ':'))
print('written', OUT)

# report stats
from collections import Counter
print('colleges union:', len(set(c['college'] for c in all_courses)))
print('categories:', len(set(c['category'] for c in all_courses)))
print('firsts:', len(set(c['first'] for c in all_courses)))
print('campuses:', dict(Counter(c['campus'] for c in all_courses)))
# autumn slots stats
slot_counts = Counter(len(c['slots']) for c in autumn_all)
print('autumn slots/course:', dict(sorted(slot_counts.items())))
